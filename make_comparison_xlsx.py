"""Generate parameter comparison Excel workbook."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
section_font = Font(bold=True, size=12, color="2F5496")
section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
calib_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
alert_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def add_headers(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = wrap
        c.border = thin_border

def add_section(ws, title, row, ncols=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=title)
    c.font = section_font
    c.fill = section_fill
    for col in range(1, ncols + 1):
        ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=col).fill = section_fill
    return row + 1

def add_row(ws, row, values, highlight_cols=None, alert_cols=None):
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.alignment = wrap
        c.border = thin_border
        if alert_cols and col in alert_cols:
            c.fill = alert_fill
        elif highlight_cols and col in highlight_cols:
            c.fill = calib_fill

# ============================================================
# SHEET 1: Main comparison
# ============================================================
ws = wb.active
ws.title = "Parameter Comparison"
headers = ["Category", "Concept", "EMOD Parameter", "EMOD Value",
           "STIsim Parameter", "STIsim Value", "STIsim Calibrated?", "Notes"]
add_headers(ws, headers)
r = 2

# TRANSMISSION
r = add_section(ws, "TRANSMISSION", r)
data = [
    ["Transmission", "Base transmission probability", "Base_Infectivity",
     "0.00233 (0.00231-0.00234) [calibrated]", "beta_m2f", "0.01",
     "YES (0.002-0.014)", "STIsim ~4x higher; EMOD compensates with larger stage multipliers"],
    ["Transmission", "Acute phase infectiousness", "Acute_Stage_Infectivity_Multiplier",
     "26x", "rel_trans_acute", "Normal(6, 0.5)x", "No",
     "EMOD 26x vs STIsim ~6x -- major structural difference"],
    ["Transmission", "Acute duration", "Acute_Duration_In_Months",
     "3 months", "dur_acute", "lognorm(mean=3mo, SD=1mo)", "No", "Match"],
    ["Transmission", "AIDS infectiousness", "AIDS_Stage_Infectivity_Multiplier",
     "4.5x", "rel_trans_falling", "Normal(8, 0.5)x", "No", "STIsim higher for late-stage"],
    ["Transmission", "AIDS duration", "AIDS_Duration_In_Months",
     "9 months", "dur_falling", "lognorm(mean=3yr, SD=1yr)", "No",
     "STIsim much longer AIDS phase"],
    ["Transmission", "F susceptibility (young <25)",
     "Male_To_Female_Relative_Infectivity_Young",
     "4.894 (4.747-5.041) [calibrated]", "rel_beta_f2m",
     "0.5 (flat, no age dependence)", "No",
     "EMOD: young women ~5x. STIsim: flat ratio. No mechanism to vary by age -- needs code change."],
    ["Transmission", "F susceptibility (old 25+)",
     "Male_To_Female_Relative_Infectivity_Old",
     "2.844 (2.727-2.958) [calibrated]", "rel_beta_f2m",
     "0.5 (same as above)", "No",
     "EMOD found significant age gradient (5x young vs 3x old)"],
    ["Transmission", "Condom efficacy", "Condom_Transmission_Blocking_Probability",
     "0.8", "eff_condom", "0.85", "YES (0.5-0.9)",
     "Efficacy when used, not probability of use"],
    ["Transmission", "Circumcision efficacy", "Circumcision_Reduced_Acquire",
     "0.6", "eff_circ", "0.6 (default)", "No", "Match"],
    ["Transmission", "STI co-infection multiplier",
     "STI_Coinfection_Acquisition_Multiplier", "5.5x", "--",
     "Not modeled", "--", "EMOD only"],
]
for vals in data:
    hl = [5, 6] if "YES" in str(vals[6]) else None
    al = [5, 6] if "No mechanism" in str(vals[7]) else None
    add_row(ws, r, vals, highlight_cols=hl, alert_cols=al)
    r += 1

# ART EFFICACY
r = add_section(ws, "ART EFFICACY & TRANSMISSION ON TREATMENT", r)
data = [
    ["ART Efficacy", "ART viral suppression", "ART_Viral_Suppression_Multiplier",
     "0.08 (92% reduction)", "art_efficacy", "0.96 (96% reduction)", "No",
     "STIsim slightly more optimistic. EMOD sensitivity showed high sensitivity (Fig A7)"],
    ["ART Efficacy", "Time to full ART effect", "--", "Immediate",
     "time_to_art_efficacy", "6 months (linear ramp)", "No",
     "STIsim ramps from 0 to 96% over 6 months"],
    ["ART Efficacy", "MTCT probability", "Maternal_Infection_Transmission_Probability",
     "0.30 per pregnancy", "beta_m2c", "0.025 per month", "No", "Different units"],
    ["ART Efficacy", "MTCT on ART multiplier", "Maternal_Transmission_ART_Multiplier",
     "0.03334", "--", "Not explicitly set", "--", ""],
]
for vals in data:
    add_row(ws, r, vals)
    r += 1

# EFFECTIVE PER-ACT
r = add_section(ws, "EFFECTIVE PER-ACT TRANSMISSION (computed)", r)
data = [
    ["Per-Act", "Chronic/latent", "--", "0.00233 x 1 = 0.0023", "--",
     "0.01 x 1 = 0.01", "--", "STIsim 4.3x higher"],
    ["Per-Act", "Acute", "--", "0.00233 x 26 = 0.061", "--",
     "0.01 x 6 = 0.06", "--", "~Match (~0.06)"],
    ["Per-Act", "AIDS", "--", "0.00233 x 4.5 = 0.010", "--",
     "0.01 x 8 = 0.08", "--", "STIsim 8x higher"],
    ["Per-Act", "On ART", "--", "0.00233 x 0.08 = 0.00019", "--",
     "0.01 x 0.04 = 0.0004", "--", "STIsim 2x higher"],
]
for vals in data:
    add_row(ws, r, vals)
    r += 1

# DISEASE PROGRESSION
r = add_section(ws, "DISEASE PROGRESSION & MORTALITY", r)
data = [
    ["Progression", "Post-infection CD4", "CD4_Post_Infection_Weibull_Scale",
     "560", "cd4_start", "Normal(800, 50)", "No", "STIsim starts higher"],
    ["Progression", "Adult survival (intercept)",
     "HIV_Adult_Survival_Scale_Parameter_Intercept", "21.182",
     "dur_latent", "lognorm(mean=10yr, SD=3yr)", "No",
     "EMOD age-dependent; STIsim fixed"],
    ["Progression", "Adult survival (age slope)",
     "HIV_Adult_Survival_Scale_Parameter_Slope", "-0.2717", "--",
     "No age dependence", "--", "Older people die faster in EMOD"],
    ["Progression", "CD4 mortality: >500", "--", "Implicit in survival curve",
     "--", "0.3% per year", "No", ""],
    ["Progression", "CD4 mortality: 350-499", "--", "Implicit in survival curve",
     "--", "0.5% per year", "No", ""],
    ["Progression", "CD4 mortality: 200-349", "--", "Implicit in survival curve",
     "--", "1% per year", "No", ""],
    ["Progression", "CD4 mortality: 50-199", "--", "Implicit in survival curve",
     "--", "5% per year", "No", ""],
    ["Progression", "CD4 mortality: <50", "--", "Implicit in survival curve",
     "--", "30% per year", "No", ""],
    ["Progression", "Child HIV (rapid progressor)",
     "HIV_Child_Survival_Rapid_Progressor_Fraction", "0.57", "--",
     "Not modeled", "--", "EMOD only"],
]
for vals in data:
    add_row(ws, r, vals)
    r += 1

# SEXUAL DEBUT
r = add_section(ws, "SEXUAL DEBUT (STIsim uses package defaults -- never overridden!)", r)
data = [
    ["Debut", "Female debut age", "Sexual_Debut_Age_Female_Weibull_Scale",
     "16.3 (16.2-16.4) [calibrated]", "debut_pars_f",
     "lognorm(mean=20, SD=3) [PACKAGE DEFAULT]", "No -- FIXED",
     "3-4 year gap! HIGH PRIORITY FIX. Override in run_sims.py."],
    ["Debut", "Male debut age", "Sexual_Debut_Age_Male_Weibull_Scale",
     "17.5 (17.4-17.7) [calibrated]", "debut_pars_m",
     "lognorm(mean=21, SD=3) [PACKAGE DEFAULT]", "No -- FIXED",
     "3-4 year gap! Delays network entry for all agents."],
    ["Debut", "Female debut shape",
     "Sexual_Debut_Age_Female_Weibull_Heterogeneity",
     "0.309 (0.293-0.322) [calibrated]", "--", "SD=3 (fixed)", "No", ""],
    ["Debut", "Male debut shape",
     "Sexual_Debut_Age_Male_Weibull_Heterogeneity",
     "0.042 (0.04-0.05) [calibrated]", "--", "SD=3 (fixed)", "No",
     "EMOD males have very tight debut distribution"],
    ["Debut", "Minimum debut age", "Sexual_Debut_Age_Min", "13", "--",
     "Not explicitly set", "--", ""],
]
for vals in data:
    al = [5, 6] if "HIGH PRIORITY" in str(vals[7]) or "3-4 year" in str(vals[7]) else None
    add_row(ws, r, vals, alert_cols=al)
    r += 1

# RISK GROUPS
r = add_section(ws, "RISK GROUPS & SEX WORK", r)
data = [
    ["Risk Groups", "Proportion low risk", "Proportion_Low_Risk",
     "0.73 (0.72-0.74) [calibrated]", "prop_f0 / prop_m0",
     "F: 0.60, M: 0.50", "YES", "EMOD single value; STIsim sex-specific"],
    ["Risk Groups", "Epidemic seed year", "Seed_Year",
     "1982.7 (1982.4-1983.2) [calibrated]", "start + init_prev",
     "1985 with CSV seeding", "No", ""],
    ["Risk Groups", "FSW proportion", "--", "3% enrollment at debut",
     "fsw_shares", "Bernoulli(0.10)", "No", "STIsim 3x higher FSW"],
    ["Risk Groups", "Client proportion", "--", "15% enrollment at debut",
     "client_shares", "Bernoulli(0.20)", "No", "STIsim higher clients"],
    ["Risk Groups", "Male MR concurrency", "--", "--", "m1_conc",
     "0.15", "YES (0.05-0.3)", ""],
]
for vals in data:
    hl = [5, 6] if "YES" in str(vals[6]) else None
    add_row(ws, r, vals, highlight_cols=hl)
    r += 1

# PARTNERSHIPS
r = add_section(ws, "PARTNERSHIP FORMATION & COITAL FREQUENCY", r)
data = [
    ["Partnerships", "Marital formation rate", "Marital_Form_Rate",
     "0.00046/day [calibrated]", "p_pair_form", "0.5 per timestep", "No",
     "Not directly comparable"],
    ["Partnerships", "Informal formation rate", "Informal_Form_Rate",
     "0.00146/day [calibrated]", "--", "Risk-group dependent", "--", ""],
    ["Partnerships", "Coital act rate", "Coital_Act_Rate",
     "0.33/day (~120/yr)", "acts", "lognorm(mean=80/yr, SD=30/yr)", "No",
     "EMOD ~50% higher"],
    ["Partnerships", "Coital dilution (2 partners)",
     "Coital_Dilution_Factor_2_Partners", "0.75", "--", "Not modeled",
     "--", "EMOD only"],
]
for vals in data:
    add_row(ws, r, vals)
    r += 1

# CONDOM USE
r = add_section(ws, "CONDOM USE PROBABILITIES (STIsim from fixed CSV, not calibrated)", r)
data = [
    ["Condom Use", "Stable/marital (max)", "Marital_Condom_Max",
     "0.218 (0.207-0.231) [calibrated]", "condom_use.csv (0,0)",
     "0.01 (2020) [FIXED CSV]", "No -- FIXED",
     "STIsim much lower for stable partnerships"],
    ["Condom Use", "Informal/HR pairs (max)", "Informal_Condom_Max",
     "0.337 (0.321-0.355) [calibrated]", "condom_use.csv (1,1)",
     "0.9 (2020) [FIXED CSV]", "No -- FIXED",
     "STIsim much higher. DHS reports ~40-60%."],
    ["Condom Use", "Transitory/cross-risk (max)", "Transitory_Condom_Max",
     "0.103 (0.089-0.117) [calibrated]", "condom_use.csv cross-risk",
     "0.7 (2020) [FIXED CSV]", "No -- FIXED", "STIsim much higher"],
    ["Condom Use", "Commercial/FSW-client (max)", "Commercial_Condom_Max",
     "0.85", "condom_use.csv (fsw,client)", "0.95 [FIXED CSV]",
     "No -- FIXED", "Similar"],
]
for vals in data:
    al = [5, 6] if "much higher" in str(vals[7]) else None
    add_row(ws, r, vals, alert_cols=al)
    r += 1

# TREATMENT CASCADE
r = add_section(ws, "TREATMENT CASCADE", r)
data = [
    ["Treatment", "ART linkage/initiation", "ART_Link_Max",
     "0.952 (0.948-0.955) [calibrated]", "art_initiation",
     "0.9 (default)", "No", "Similar (~90-95%)"],
    ["Treatment", "ART linkage midpoint year", "ART_Link_Mid",
     "2010.7 (2010.4-2010.9) [calibrated]", "--",
     "Immediate after diagnosis", "--", "EMOD time-varying linkage"],
    ["Treatment", "ART dropout (base duration)", "ART_dropout",
     "7300 days (~20yr mean)", "dur_on_art",
     "lognorm(mean=3yr, SD=1.5yr)", "No (base)",
     "Base differs 7x, but see scaling below"],
    ["Treatment", "ART duration scaling", "--", "--", "rel_dur_on_art",
     "guess=8 (range 1-20)", "YES",
     "At 8x: effective ~24yr, close to EMOD 20yr"],
    ["Treatment", "Delay infection to ART", "Delay_Period_Mean",
     "180 days", "--", "Immediate after diagnosis", "--",
     "EMOD has explicit delay"],
    ["Treatment", "ART coverage format", "--",
     "Campaign (absolute counts)", "--",
     "Proportion by age/sex/year (CSV)", "--", "STIsim now stratified"],
    ["Treatment", "FSW testing (2020)", "--", "Campaign-based", "--",
     "75% annual probability", "No", ""],
    ["Treatment", "General pop testing (2020)", "--", "Campaign-based",
     "--", "50% annual probability", "No", ""],
    ["Treatment", "Low-CD4 testing (2020)", "--", "Symptomatic cascade",
     "--", "85% annual (CD4<200)", "No", ""],
    ["Treatment", "PrEP efficacy", "Waning function",
     "50% at day 0, 0% at 365", "eff_prep", "0.80 (constant)", "No",
     "EMOD wanes; STIsim constant"],
    ["Treatment", "VMMC status", "--", "Active (campaign)", "--",
     "Commented out", "--", "STIsim not using VMMC"],
]
for vals in data:
    hl = [5, 6] if "YES" in str(vals[6]) else None
    add_row(ws, r, vals, highlight_cols=hl)
    r += 1

# INITIAL PREVALENCE
r = add_section(ws, "INITIAL PREVALENCE & EPIDEMIC SEEDING", r)
data = [
    ["Init Prev", "Epidemic seed", "Seed_Year",
     "1982.7 [calibrated]", "sim start",
     "1985 + init_prev CSV", "No",
     "EMOD sim starts 1960, seeds 1983; STIsim starts 1985"],
    ["Init Prev", "rel_init_prev scaling", "--", "--", "rel_init_prev",
     "0.1 (was calibrated, now FIXED)", "No (commented out)",
     "All CSV values multiplied by 0.1"],
    ["Init Prev", "LR Female, non-FSW", "--", "--", "--",
     "0.1% x 0.1 = 0.01%", "No", ""],
    ["Init Prev", "MR Female, non-FSW", "--", "--", "--",
     "1% x 0.1 = 0.1%", "No", ""],
    ["Init Prev", "HR Female, non-FSW", "--", "--", "--",
     "5% x 0.1 = 0.5%", "No", ""],
    ["Init Prev", "FSW (any risk)", "--", "--", "--",
     "15-20% x 0.1 = 1.5-2.0%", "No", ""],
    ["Init Prev", "LR Male, non-client", "--", "--", "--",
     "0.1% x 0.1 = 0.01%", "No", ""],
    ["Init Prev", "MR Male, non-client", "--", "--", "--",
     "1% x 0.1 = 0.1%", "No", ""],
    ["Init Prev", "HR Male, non-client", "--", "--", "--",
     "2% x 0.1 = 0.2%", "No", ""],
    ["Init Prev", "Male clients", "--", "--", "--",
     "5-15% x 0.1 = 0.5-1.5%", "No", ""],
    ["Init Prev", "Who is seeded?", "--",
     "High-risk groups at Seed_Year", "--",
     "Agents past debut age at sim start", "--",
     "Late debut (20yr) means many 15-19yr olds NOT seeded"],
    ["Init Prev", "Infection backdating", "--", "--",
     "dist_ti_init_infected", "Uniform(-120mo, -5mo)", "No",
     "Agents start at various disease stages"],
]
for vals in data:
    add_row(ws, r, vals)
    r += 1

# Column widths
for i, w in enumerate([14, 30, 40, 36, 26, 36, 20, 58], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# ============================================================
# SHEET 2: Calibration summary
# ============================================================
ws2 = wb.create_sheet("Calibration Parameters")
headers2 = ["Model", "Parameter", "Category", "Low", "High",
            "Guess/Median (IQR)", "Notes"]
add_headers(ws2, headers2)
r = 2

r = add_section(ws2, "STIsim Calibrated Parameters (6 total)", r, ncols=7)
for vals in [
    ["STIsim", "beta_m2f", "Transmission", "0.002", "0.014", "0.006",
     "Male-to-female per-act probability"],
    ["STIsim", "eff_condom", "Transmission", "0.5", "0.9", "0.75",
     "Condom efficacy (NOT condom use probability)"],
    ["STIsim", "rel_dur_on_art", "Treatment", "1.0", "20.0", "8.0",
     "Multiplier on base 3yr ART duration"],
    ["STIsim", "prop_f0", "Network", "0.55", "0.9", "0.7",
     "Proportion females low-risk"],
    ["STIsim", "prop_m0", "Network", "0.55", "0.8", "0.65",
     "Proportion males low-risk"],
    ["STIsim", "m1_conc", "Network", "0.05", "0.3", "0.15",
     "Male medium-risk concurrency"],
]:
    add_row(ws2, r, vals, highlight_cols=[2, 4, 5, 6])
    r += 1

r = add_section(ws2, "EMOD Calibrated Parameters (24 total -- Lancet HIV 2020 Table A2)", r, ncols=7)
for vals in [
    ["EMOD", "Base_Infectivity", "Transmission", "", "",
     "0.00233 (0.00231-0.00234)", ""],
    ["EMOD", "Male_To_Female_Relative_Infectivity_Young", "Transmission",
     "", "", "4.894 (4.747-5.041)", "Age <25"],
    ["EMOD", "Male_To_Female_Relative_Infectivity_Old", "Transmission",
     "", "", "2.844 (2.727-2.958)", "Age 25+"],
    ["EMOD", "Sexual_Debut_Age_Female_Weibull_Scale", "Network", "", "",
     "16.302 (16.166-16.396)", ""],
    ["EMOD", "Sexual_Debut_Age_Female_Weibull_Heterogeneity", "Network",
     "", "", "0.309 (0.293-0.322)", ""],
    ["EMOD", "Sexual_Debut_Age_Male_Weibull_Scale", "Network", "", "",
     "17.499 (17.357-17.699)", ""],
    ["EMOD", "Sexual_Debut_Age_Male_Weibull_Heterogeneity", "Network",
     "", "", "0.042 (0.04-0.05)", ""],
    ["EMOD", "Proportion_Low_Risk", "Network", "", "",
     "0.73 (0.721-0.742)", ""],
    ["EMOD", "Seed_Year", "Init Conditions", "", "",
     "1982.7 (1982.4-1983.2)", ""],
    ["EMOD", "Marital_Form_Rate", "Network", "", "",
     "0.00046 (0.00044-0.0005)", "per day"],
    ["EMOD", "Informal_Form_Rate", "Network", "", "",
     "0.00146 (0.00134-0.00155)", "per day"],
    ["EMOD", "Marital_Condom_Max", "Condom Use", "", "",
     "0.218 (0.207-0.231)", ""],
    ["EMOD", "Marital_Condom_Mid", "Condom Use", "", "",
     "2001.8 (2001.5-2002.1)", "year"],
    ["EMOD", "Marital_Condom_Rate", "Condom Use", "", "",
     "2.407 (2.252-2.524)", ""],
    ["EMOD", "Informal_Condom_Max", "Condom Use", "", "",
     "0.337 (0.321-0.355)", ""],
    ["EMOD", "Informal_Condom_Mid", "Condom Use", "", "",
     "1992.6 (1992.2-1992.9)", "year"],
    ["EMOD", "Informal_Condom_Rate", "Condom Use", "", "",
     "3.003 (2.941-3.076)", ""],
    ["EMOD", "Transitory_Condom_Max", "Condom Use", "", "",
     "0.103 (0.089-0.117)", ""],
    ["EMOD", "Transitory_Condom_Mid", "Condom Use", "", "",
     "1996.7 (1996.1-1997)", "year"],
    ["EMOD", "Transitory_Condom_Rate", "Condom Use", "", "",
     "2.998 (2.878-3.106)", ""],
    ["EMOD", "ART_Link_Max", "Treatment", "", "",
     "0.952 (0.948-0.955)", ""],
    ["EMOD", "ART_Link_Mid", "Treatment", "", "",
     "2010.7 (2010.4-2010.9)", "year"],
    ["EMOD", "preART_Link_Max", "Treatment", "", "",
     "0.807 (0.783-0.829)", ""],
    ["EMOD", "preART_Link_Mid", "Treatment", "", "",
     "1995.7 (1995.1-1996.4)", "year"],
]:
    add_row(ws2, r, vals, highlight_cols=[2, 6])
    r += 1

for i, w in enumerate([10, 48, 15, 8, 8, 34, 42], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"

# ============================================================
# SHEET 3: Suggested changes
# ============================================================
ws3 = wb.create_sheet("Suggested Changes")
headers3 = ["Priority", "Area", "Current STIsim Value",
            "Suggested Change", "Rationale", "Difficulty"]
add_headers(ws3, headers3)
r = 2

for vals in [
    ["1 - HIGH", "Sexual debut age",
     "F: lognorm(20,3), M: lognorm(21,3)",
     "Override to F:[16.3, 2], M:[17.5, 1.5] in run_sims.py",
     "Package default, not country-specific. EMOD calibrated ~16-17yr. DHS data supports earlier debut.",
     "Easy - 2 lines in run_sims.py"],
    ["2 - HIGH", "Condom use probabilities",
     "70-95% for non-stable (fixed CSV)",
     "Lower to DHS levels (~40-60%) or add to calibration",
     "Current values likely unrealistic. EMOD calibrated 10-34%. DHS/PHIA report 40-60%.",
     "Easy - edit condom_use.csv"],
    ["3 - MEDIUM", "Age-dependent F susceptibility",
     "Flat rel_beta_f2m=0.5",
     "Implement age-varying susceptibility (5x young, 3x old)",
     "EMOD calibrated to 4.9x (<25) vs 2.8x (25+). Important for age-specific incidence.",
     "Hard - requires stisim code change"],
    ["4 - MEDIUM", "ART efficacy",
     "0.96 (96% reduction)",
     "Lower to 0.92 to match EMOD and Donnell et al.",
     "EMOD sensitivity analysis showed high sensitivity. 4% difference matters.",
     "Easy - 1 line in run_sims.py"],
    ["5 - MEDIUM", "Add calibration parameters",
     "Only 6 free parameters",
     "Add debut age, condom use, rel_beta_f2m, rel_init_prev",
     "EMOD uses 24 parameters. More freedom = better fit to age-specific data.",
     "Medium - edit run_calibrations.py"],
    ["6 - LOW", "Initial prevalence scaling",
     "rel_init_prev=0.1 (fixed)",
     "Re-enable as calibration parameter",
     "Was previously calibrated, now commented out. Affects epidemic take-off.",
     "Easy - uncomment in run_calibrations.py"],
]:
    al = [1] if "HIGH" in vals[0] else None
    add_row(ws3, r, vals, alert_cols=al)
    r += 1

for i, w in enumerate([14, 30, 38, 48, 60, 32], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = "A2"

# Save
outpath = "parameter_comparison_EMOD_vs_STIsim.xlsx"
wb.save(outpath)
print(f"Saved {outpath}")
